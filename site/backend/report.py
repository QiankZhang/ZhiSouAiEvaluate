"""内置报告生成器（evaluation-report skill 的平台内置实现）。

不重新打分，消费任务的逐条评测结果，产出：
- 结构化 content（供前端 ReportView 渲染）
- 五段式 Markdown 总报告（整体结论 → GSB 专项 → 分维度问题分析 → 典型错误 case → 改进建议）
- Excel 原始打分表（含 GSB 汇总 sheet），逻辑复刻 skills/evaluation-report/scripts/make_scores_xlsx.py
"""

import io
import json
import logging
from typing import Any

from . import config, llm

logger = logging.getLogger(__name__)

LOW_SCORE_THRESHOLD = 3.0  # evaluation-report SKILL.md 默认低分阈值
DIM_LOW_MARK = 2  # 维度 score <= 2 记为该维度低分

# 人工复核状态的中文展示标签（与前端 api.js 的 REVIEW_LABELS / RESULT_REVIEW_LABELS 一致）。
# 定义在这里而非 main.py：report.py 是唯一消费方，放这里可去掉 report→main 的延迟 import。
REVIEW_STATUS_LABELS = {"NOT_STARTED": "未开始", "IN_PROGRESS": "复核中", "COMPLETED": "已复核"}
RESULT_REVIEW_STATUS_LABELS = {"PENDING": "待复核", "APPROVED": "已通过", "ADJUSTED": "已调整"}


def _effective_scores(r: dict[str, Any]) -> dict[str, Any]:
    """人工复核调整过的分数优先。"""
    base = r.get("scores") or {}
    adj = r.get("adjusted_scores")
    return {**base, **adj} if adj else base


# ---------- 典型错误 case 分析（对应 evaluation-report SKILL.md 第四节 4.1/4.2） ----------

ERROR_TYPES = ["数据错误", "数据过时", "数据缺失", "数据呈现", "触发风控", "其他"]


def _analyze_badcases_llm(badcases: list[dict[str, Any]], model: str) -> dict[int, dict[str, str]] | None:
    """对每个错误 case 定位错误类型 + 从原文摘取问题片段（逐字引用，不是复述）。"""
    if not badcases:
        return None
    payload = [
        {"index": i, "query": b["query"], "content": (b.get("content") or "")[:600], "reason": b["reason"]}
        | ({"baseline": b["baseline"][:400]} if b.get("baseline") else {})
        for i, b in enumerate(badcases[:30])
    ]
    try:
        resp = llm.chat_completion(
            [
                {
                    "role": "system",
                    "content": (
                        "你是评估报告分析助手。只输出一个 JSON 对象，形如："
                        '{"cases": [{"index": <整数>, "error_type": "<' + "/".join(ERROR_TYPES) + '>", '
                        '"problem_span": "<content 原文逐字摘录，≤80字>", "insight": "<一句话根因，≤60字>"}]}'
                        "，不要输出任何解释或代码块。"
                    ),
                },
                {
                    "role": "user",
                    "content": (
                        "以下是低分/劣于基线的 case（下标从 0 开始）。请为每条：\n"
                        "1) 归属错误类型（数据错误/数据过时/数据缺失/数据呈现/触发风控/其他）；\n"
                        "2) 从 content 原文中摘取一段能体现问题的原文片段（problem_span 必须逐字引用原文，不要复述改写）；\n"
                        "3) 用一句话点出根因（insight）。\n"
                        + json.dumps(payload, ensure_ascii=False, indent=1)
                    ),
                },
            ],
            model=model,
            response_format={"type": "json_object"},
        )
        args = llm.parse_json_response(resp)
    except (llm.LlmError, ValueError, json.JSONDecodeError):
        return None

    out: dict[int, dict[str, str]] = {}
    for c in args.get("cases", []):
        i = c.get("index")
        if isinstance(i, int) and 0 <= i < len(badcases):
            out[i] = {
                "error_type": c.get("error_type") if c.get("error_type") in ERROR_TYPES else "其他",
                "problem_span": (c.get("problem_span") or "").strip(),
                "insight": (c.get("insight") or "").strip(),
            }
    return out or None


def _analyze_badcases_rule(badcases: list[dict[str, Any]]) -> dict[int, dict[str, str]]:
    """未启用 LLM（模拟引擎）时的兜底：错误类型统一记为"其他"，问题片段退化为原文开头节选。"""
    out: dict[int, dict[str, str]] = {}
    for i, b in enumerate(badcases):
        content = (b.get("content") or "").strip()
        span = content if len(content) <= 50 else content[:50] + "…"
        out[i] = {"error_type": "其他", "problem_span": span, "insight": ""}
    return out


def _pick_typical_cases(badcases: list[dict[str, Any]], limit: int = 3) -> list[dict[str, Any]]:
    """优先覆盖不同错误类型选 `limit` 个典型 case，不够再按顺序补齐（对应 SKILL.md 4.2）。"""
    seen_types: set[str] = set()
    picked_idx: list[int] = []
    for i, b in enumerate(badcases):
        t = b.get("error_type") or "其他"
        if t not in seen_types:
            picked_idx.append(i)
            seen_types.add(t)
        if len(picked_idx) >= limit:
            break
    if len(picked_idx) < limit:
        for i in range(len(badcases)):
            if i not in picked_idx:
                picked_idx.append(i)
            if len(picked_idx) >= limit:
                break
    return [badcases[i] for i in picked_idx]


def _build_error_cases(badcases: list[dict[str, Any]], model: str, use_llm: bool) -> dict[str, Any]:
    if not badcases:
        return {"total": 0, "type_counts": [], "typical": []}
    analysis = (use_llm and _analyze_badcases_llm(badcases, model)) or _analyze_badcases_rule(badcases)
    for i, b in enumerate(badcases):
        b.update(analysis.get(i, {"error_type": "其他", "problem_span": "", "insight": ""}))
    counts: dict[str, int] = {}
    for b in badcases:
        counts[b["error_type"]] = counts.get(b["error_type"], 0) + 1
    type_counts = [
        {"type": t, "count": c, "ratio": round(c / len(badcases) * 100, 1)}
        for t, c in sorted(counts.items(), key=lambda x: -x[1])
    ]
    typical = [
        {
            "query": b["query"],
            "content": b.get("content", ""),
            "baseline": b.get("baseline") or None,
            "error_type": b["error_type"],
            "problem_span": b["problem_span"],
            "insight": b["insight"] or b["reason"],
        }
        for b in _pick_typical_cases(badcases)
    ]
    return {"total": len(badcases), "type_counts": type_counts, "typical": typical}


# ---------- 结构化 content ----------

def build_report(task: dict[str, Any], results: list[dict[str, Any]], benchmark: dict[str, Any]) -> dict[str, Any]:
    eval_method = benchmark["eval_method"]
    total = len(results)
    if total == 0:
        return {"status": "READY", "eval_method": eval_method, "content": {"summary": {"total": 0}}}

    use_llm = config.engine_for(task["judge_model"]) == "agent" and llm.is_live(task["judge_model"])

    if eval_method == "GSB":
        good = sum(1 for r in results if _effective_scores(r).get("judgment") == "Good")
        same = sum(1 for r in results if _effective_scores(r).get("judgment") == "Same")
        bad = sum(1 for r in results if _effective_scores(r).get("judgment") == "Bad")
        badcases = [
            {"query": r["query"], "content": r["content"], "baseline": r.get("baseline", ""), "reason": r["reason"]}
            for r in results
            if _effective_scores(r).get("judgment") == "Bad"
        ]
        denom = good + same + bad
        error_cases = _build_error_cases(badcases, task["judge_model"], use_llm)
        content = {
            "summary": {
                "total": total,
                "good": good,
                "same": same,
                "bad": bad,
                "win_rate": round(good / total * 100, 1),
                "net_win_rate": round((good - bad) / denom * 100, 1) if denom else 0.0,
            },
            "error_cases": error_cases,
            "suggestions": _suggestions(eval_method, None, error_cases["typical"]),
        }
        return {"status": "READY", "eval_method": eval_method, "content": content}

    dims = benchmark["config"]["dimensions"]
    dim_stats = []
    for d in dims:
        vals, lows = [], 0
        for r in results:
            for sd in _effective_scores(r).get("dimensions", []):
                if sd.get("key") == d["key"]:
                    vals.append(sd["score"])
                    lows += 1 if sd["score"] <= DIM_LOW_MARK else 0
        avg = round(sum(vals) / len(vals), 2) if vals else 0.0
        dim_stats.append(
            {
                "key": d["key"],
                "name": d.get("name", d["key"]),
                "avg": avg,
                "weight": d.get("weight", 0),
                "low_count": lows,
                "low_ratio": round(lows / total * 100, 1),
            }
        )
    totals = [_effective_scores(r)["total"] for r in results]
    confs = [r["confidence"] for r in results if r.get("confidence") is not None]
    low_results = [r for r in results if _effective_scores(r)["total"] < LOW_SCORE_THRESHOLD]
    badcases = [{"query": r["query"], "content": r["content"], "reason": r["reason"]} for r in low_results]
    weakest = min(dim_stats, key=lambda x: x["avg"]) if dim_stats else None
    strongest = max(dim_stats, key=lambda x: x["avg"]) if dim_stats else None
    error_cases = _build_error_cases(badcases, task["judge_model"], use_llm)
    content = {
        "summary": {
            "total": total,
            "avg_total": round(sum(totals) / len(totals), 2),
            "avg_confidence": round(sum(confs) / len(confs), 2) if confs else None,
            "low_count": len(low_results),
            "low_ratio": round(len(low_results) / total * 100, 1),
            "weakest_dim": weakest["name"] if weakest else None,
            "strongest_dim": strongest["name"] if strongest else None,
        },
        "dimensions": dim_stats,
        "distribution": {str(s): sum(1 for v in totals if round(v) == s) for s in range(1, 6)},
        "error_cases": error_cases,
        "suggestions": _suggestions(eval_method, weakest["name"] if weakest else None, error_cases["typical"]),
    }
    return {"status": "READY", "eval_method": eval_method, "content": content}


def _suggestions(eval_method: str, weakest_dim: str | None, typical_cases: list[dict[str, Any]]) -> list[str]:
    """每条建议尽量关联一个具体 case 作为依据，避免空泛（对应 SKILL.md 第五节要求）。"""
    cite = None
    if typical_cases:
        ex = typical_cases[0]
        span = ex.get("problem_span") or ex.get("insight") or ""
        cite = f"如案例「{ex['query']}」（{ex['error_type']}）：{span}"

    if eval_method == "GSB":
        tips = ["对净胜率为负的问题分类，优先补齐对应的召回与内容整合策略。"]
        if cite:
            tips.append(f"{cite}——建议针对此类问题补充人工归因规则，沉淀为回归用例集。")
        else:
            tips.append("针对被判 Bad 的样本做人工归因，沉淀为回归用例集。")
        return tips

    tips = [
        "召回阶段优先引入官方 / 权威站点，提升准确性与时效性。",
        "对含数据结论的结果强制来源标注，降低数据错误与过时风险。",
    ]
    if cite:
        tips.insert(0, f"{cite}——建议优先修复此类问题。")
    if weakest_dim:
        tips.insert(0 if not cite else 1, f"「{weakest_dim}」为当前最弱维度，建议作为下一轮策略迭代的重点。")
    return tips


# ---------- 报告模板驱动的 LLM 报告生成 ----------

GSB_SECTION = "GSB 专项评估"  # 仅当任务为 GSB 对比时才纳入报告

# 「提示词」类型报告模板的默认章节，用户可在报告模板里增删。
DEFAULT_REPORT_SECTIONS = [
    "整体结论",
    GSB_SECTION,
    "分维度问题分析",
    "典型错误 case 分析",
    "改进建议",
]

# 「提示词」类型报告模板的默认提示词。{章节} 会被替换为模板配置的章节清单。
DEFAULT_REPORT_PROMPT = (
    "你是「智搜策略效果评估」平台的评估报告撰写助手。基于用户提供的评测统计与样本数据，"
    "撰写一份 Markdown 格式的评估总报告。\n\n"
    "报告需覆盖以下章节（先后顺序、标题措辞可自行安排，但每一项内容都要覆盖到）：\n"
    "{章节}\n\n"
    "要求：结论先行；每个论断都用具体数字或样本 query 佐证，不写空泛套话；层次清晰；"
    "开头用一级标题「# <任务名> · 评估报告」并列出关键元信息。"
)


def _template_sections(cfg: dict[str, Any], is_gsb: bool) -> list[str]:
    raw = cfg.get("sections") or DEFAULT_REPORT_SECTIONS
    out = [s.strip() for s in raw if isinstance(s, str) and s.strip()]
    if not is_gsb:
        out = [s for s in out if s != GSB_SECTION]
    fallback = [s for s in DEFAULT_REPORT_SECTIONS if is_gsb or s != GSB_SECTION]
    return out or fallback


def _report_payload(
    task: dict[str, Any], report: dict[str, Any], results: list[dict[str, Any]], is_gsb: bool
) -> dict[str, Any]:
    content = report.get("content", {})
    # 只送精简明细：query + 分数/判定 + 简短理由。原文片段仅错误 case 需要，已在
    # content["error_cases"] 里携带——整份 content 原文会让提示词膨胀到数万 token 而超时。
    samples: list[dict[str, Any]] = []
    for r in results[:120]:
        s = _effective_scores(r)
        row: dict[str, Any] = {"query": r["query"], "reason": (r.get("reason") or "")[:160]}
        if is_gsb:
            row["judgment"] = s.get("judgment")
            row["exp_vs_base"] = f"{s.get('total')} vs {s.get('baseline_total')}"
        else:
            row["total"] = s.get("total")
            row["dimensions"] = {d.get("name", d.get("key")): d.get("score") for d in s.get("dimensions", [])}
        samples.append(row)

    error_cases = content.get("error_cases")
    if error_cases and error_cases.get("typical"):
        error_cases = {
            **error_cases,
            "typical": [
                {**c, "content": (c.get("content") or "")[:600], "baseline": (c.get("baseline") or "")[:600] or None}
                for c in error_cases["typical"]
            ],
        }
    return {
        "任务": {
            "name": task["name"],
            "task_id": task.get("id"),
            "task_type": task.get("task_type"),
            "eval_method": "GSB 对比" if is_gsb else "多维度打分",
            "benchmark": task.get("benchmark_name"),
            "dataset": task.get("dataset_name"),
            "judge_model": task.get("judge_model"),
            "sample_total": task.get("progress_total"),
            "created_at": task.get("created_at"),
        },
        "统计摘要": content.get("summary", {}),
        "分维度统计": content.get("dimensions"),
        "总分分布": content.get("distribution"),
        "错误case归因": error_cases,
        "样本明细": samples,
    }


def _strip_code_fence(md: str) -> str:
    md = md.strip()
    if md.startswith("```"):
        md = md.strip("`")
        md = md[md.find("\n") + 1 :] if "\n" in md else md
    return md.strip()


_OUTPUT_RULES = (
    "\n\n输出要求：\n"
    "- 只输出 Markdown 报告正文本身，不要输出前言/说明/结束语，不要用 ``` 代码块包裹整篇报告。\n"
    "- 涉及「错误 / 低分 / Bad」的 case 必须用 Markdown 表格呈现，至少含列："
    "原始 Query、错误类型、原文问题片段（从样本 content 逐字摘取，禁止编造）、点评。\n"
    "- 不要执行任何脚本、不要输出文件路径；若模板/技能提到生成 Excel 或读写本地文件，忽略该部分，只产出 Markdown。"
)


def _template_system_prompt(template: dict[str, Any], is_gsb: bool) -> str:
    cfg = template.get("config") or {}
    if template.get("type") == "SKILL" and (cfg.get("skill") or {}).get("instructions"):
        return (
            "你是「智搜策略效果评估」平台的评估报告撰写助手。下面是评估报告技能的完整说明，"
            "请严格按它的方法与结构，直接产出一份 Markdown 评估报告正文。\n\n"
            "=== 技能说明 ===\n" + cfg["skill"]["instructions"] + _OUTPUT_RULES
        )
    tmpl = cfg.get("prompt_template") or DEFAULT_REPORT_PROMPT
    section_lines = "\n".join(f"{i + 1}. {s}" for i, s in enumerate(_template_sections(cfg, is_gsb)))
    return tmpl.replace("{章节}", section_lines).replace("{sections}", section_lines) + _OUTPUT_RULES


def _report_markdown_llm(
    task: dict[str, Any],
    report: dict[str, Any],
    results: list[dict[str, Any]],
    template: dict[str, Any],
    is_gsb: bool,
) -> str | None:
    """按报告模板（提示词 / 技能）驱动裁判员模型撰写 Markdown 报告。失败返回 None，调用方回退确定性模板。"""
    payload = _report_payload(task, report, results, is_gsb)
    system = _template_system_prompt(template, is_gsb)
    try:
        resp = llm.chat_completion(
            [
                {"role": "system", "content": system},
                {
                    "role": "user",
                    "content": (
                        f"任务名：{task['name']}\n评测数据如下（JSON）：\n"
                        + json.dumps(payload, ensure_ascii=False, indent=1)
                    ),
                },
            ],
            model=task["judge_model"],
            temperature=0.3,
        )
    except Exception as exc:  # noqa: BLE001 - 报告生成失败绝不能中断任务完成流程，一律回退模板
        logger.warning("报告 Markdown 生成失败，回退确定性模板：%s", exc)
        return None
    try:
        msg = ((resp.get("choices") or [{}])[0].get("message") or {}).get("content") or ""
        md = _strip_code_fence(msg)
    except Exception as exc:  # noqa: BLE001
        logger.warning("报告 Markdown 解析失败，回退确定性模板：%s", exc)
        return None
    if not md:
        logger.warning("报告 Markdown 生成返回空内容，回退确定性模板")
    return md or None


def generate_report_markdown(
    task: dict[str, Any],
    report: dict[str, Any],
    results: list[dict[str, Any]],
    template: dict[str, Any] | None = None,
) -> str:
    """任务完成时产出报告 Markdown：按所选报告模板走真实模型生成，无模板/模拟引擎/失败时回退确定性五段式模板。"""
    is_gsb = report.get("eval_method") == "GSB"
    use_llm = config.engine_for(task["judge_model"]) == "agent" and llm.is_live(task["judge_model"])
    if use_llm and template:
        md = _report_markdown_llm(task, report, results, template, is_gsb)
        if md:
            return md
    return report_to_markdown(task, report, results)


# ---------- Markdown（确定性五段式模板，作为回退） ----------

def _md_cell(text: Any) -> str:
    return str(text).replace("|", "\\|").replace("\n", " ").strip()


def report_to_markdown(task: dict[str, Any], report: dict[str, Any], results: list[dict[str, Any]]) -> str:
    content = report.get("content", {})
    summary = content.get("summary", {})
    is_gsb = report.get("eval_method") == "GSB"
    review_label = REVIEW_STATUS_LABELS.get(task["review_status"], task["review_status"])
    method_label = "GSB 对比" if is_gsb else "多维度"

    L: list[str] = [
        f"# {task['name']} · 评估报告",
        "",
        f"- **任务 ID**：`{task['id']}`　**任务类型**：{task['task_type']}　**评估方式**：{method_label}",
        f"- **评估基准**：{task['benchmark_name']}　**评测数据**：{task['dataset_name']}",
        f"- **裁判员模型**：{task['judge_model']}　**评测引擎**：{task.get('engine', 'simulated')}"
        + ("（部分降级为模拟）" if task.get("engine_downgraded") else ""),
        f"- **样本量**：{task['progress_total']}　**人工复核状态**：{review_label}　**创建时间**：{task['created_at']}",
        "",
        "## 一、整体结论",
        "",
    ]
    if is_gsb:
        L.append(
            f"- **GSB 速览**：G:S:B = {summary.get('good', 0)}:{summary.get('same', 0)}:{summary.get('bad', 0)}"
            f"，胜率 {summary.get('win_rate', 0)}%，净胜率 {summary.get('net_win_rate', 0)}%。"
        )
        verdict = "整体弱于基线" if summary.get("net_win_rate", 0) < 0 else "整体持平或占优基线"
        L += [f"- **核心结论**：{verdict}；被判 Bad 的样本集中体现召回与整合环节的短板。", ""]
        L += [
            "## 二、GSB 专项评估",
            "",
            "| Good | Same | Bad | 胜率 | 净胜率 |",
            "| --- | --- | --- | --- | --- |",
            f"| {summary.get('good', 0)} | {summary.get('same', 0)} | {summary.get('bad', 0)} "
            f"| {summary.get('win_rate', 0)}% | {summary.get('net_win_rate', 0)}% |",
            "",
            "> 净胜率 = (G − B) / (G + S + B)",
            "",
        ]
    else:
        L += [
            f"- **规模与低分**：总样本 {summary.get('total', 0)}，低分样本 {summary.get('low_count', 0)} 个、"
            f"占比 {summary.get('low_ratio', 0)}%（总分 < {LOW_SCORE_THRESHOLD}）。",
            f"- **核心结论**：最弱维度为「{summary.get('weakest_dim', '—')}」，最强维度为「{summary.get('strongest_dim', '—')}」；"
            "问题贯穿需求理解—物料获取—内容整合—结果呈现全链路。",
            f"- **平均总分**：{summary.get('avg_total', '—')}　**平均置信度**：{summary.get('avg_confidence', '—')}",
            "",
            "## 三、分维度问题分析",
            "",
            "| 维度 | 平均分 | 权重 | 低分数(≤2) | 低分占比 |",
            "| --- | --- | --- | --- | --- |",
        ]
        for d in content.get("dimensions", []):
            L.append(
                f"| {d['name']} | {d['avg']} | {d['weight']}% | {d['low_count']} | {d['low_ratio']}% |"
            )
        L.append("")

    error_cases = content.get("error_cases") or {"total": 0, "type_counts": [], "typical": []}
    L += ["## 四、典型错误 case 分析", ""]
    if error_cases["total"]:
        L += ["### 4.1 错误归因聚合", ""]
        agg = "、".join(f"{t['type']} {t['count']} 个（{t['ratio']}%）" for t in error_cases["type_counts"])
        L += [f"合计 {error_cases['total']} 个：{agg}。", "", "### 4.2 典型错误 case 详评", ""]
        for ex in error_cases["typical"]:
            L.append(f"- **原始 Query**：{_md_cell(ex['query'])}")
            L.append(f"- **错误定位**：{ex['error_type']}")
            if ex.get("content"):
                L.append(f"- **智搜回答**（节选）：{_md_cell(ex['content'])[:120]}")
            L.append(f"- **原文中的问题点**：「{_md_cell(ex['problem_span'])}」" if ex.get("problem_span") else "- **原文中的问题点**：—")
            if ex.get("baseline"):
                L.append(f"- **竞品表现**（节选）：{_md_cell(ex['baseline'])[:120]}")
            L.append(f"- **核心点评**：{_md_cell(ex['insight'])}")
            L.append("")
    else:
        L += ["无低分 / Bad 样本。", ""]

    L += ["## 五、改进建议", ""]
    for s in content.get("suggestions", []):
        L.append(f"- {s}")
    L.append("")

    if results:
        L += ["## 附：逐样本明细", ""]
        if is_gsb:
            L += ["| # | Query | 判定 | 实验 vs 基线 | 理由 |", "| --- | --- | --- | --- | --- |"]
            for r in results:
                s = _effective_scores(r)
                L.append(
                    f"| {r['row_index']} | {_md_cell(r['query'])} | {s.get('judgment', '—')} "
                    f"| {s.get('total', '—')} vs {s.get('baseline_total', '—')} | {_md_cell(r['reason'])} |"
                )
        else:
            L += ["| # | Query | 总分 | 维度得分 | 理由 |", "| --- | --- | --- | --- | --- |"]
            for r in results:
                s = _effective_scores(r)
                dim_str = "、".join(f"{d['name']} {d['score']}" for d in s.get("dimensions", []))
                L.append(
                    f"| {r['row_index']} | {_md_cell(r['query'])} | {s.get('total', '—')} | {dim_str} | {_md_cell(r['reason'])} |"
                )
        L.append("")

    return "\n".join(L)


# ---------- Excel ----------

def report_to_xlsx(task: dict[str, Any], results: list[dict[str, Any]], benchmark: dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    is_gsb = benchmark["eval_method"] == "GSB"
    dims = benchmark["config"]["dimensions"]
    dim_names = [d.get("name", d["key"]) for d in dims]

    wb = Workbook()
    ws = wb.active
    ws.title = "原始打分表"
    header = ["排名", "条目ID", "Query"] + dim_names + ["总分", "理由", "复核状态"]
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDEBF7")
        c.alignment = Alignment(horizontal="center")

    ranked = sorted(results, key=lambda r: -float(_effective_scores(r).get("total") or 0))
    for i, r in enumerate(ranked, 1):
        s = _effective_scores(r)
        by_key = {d.get("key"): d.get("score") for d in s.get("dimensions", [])}
        row = [i, f"item-{r['row_index']}", r["query"]]
        row += [by_key.get(d["key"]) for d in dims]
        row += [
            s.get("total"),
            r["reason"],
            RESULT_REVIEW_STATUS_LABELS.get(r["review_status"], r["review_status"]),
        ]
        ws.append(row)
    ws.freeze_panes = "A2"
    for idx, w in enumerate([6, 12, 30] + [10] * len(dim_names) + [8, 40, 10], 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w

    if is_gsb:
        gs = wb.create_sheet("GSB汇总")
        gs.append(["分类", "G", "S", "B", "净胜率"])
        for c in gs[1]:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="FCE4D6")
            c.alignment = Alignment(horizontal="center")
        g = sum(1 for r in results if _effective_scores(r).get("judgment") == "Good")
        s_ = sum(1 for r in results if _effective_scores(r).get("judgment") == "Same")
        b = sum(1 for r in results if _effective_scores(r).get("judgment") == "Bad")
        denom = g + s_ + b
        net = f"{(g - b) / denom * 100:+.1f}%" if denom else "—"
        gs.append(["整体", g, s_, b, net])
        gs.append([])
        gs.append(["逐 case", "row_index", "judgment", "实验 vs 基线", "理由"])
        for c in gs[gs.max_row]:
            c.font = Font(bold=True, italic=True)
        for r in results:
            sc = _effective_scores(r)
            gs.append(["", r["row_index"], sc.get("judgment", ""), f"{sc.get('total')} vs {sc.get('baseline_total')}", r["reason"]])
        for col, w in zip("ABCDE", [10, 12, 12, 16, 48]):
            gs.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
