#!/usr/bin/env python3
"""make_scores_xlsx.py — 将多维度评测结果(JSON)转为原始打分表 Excel(.xlsx)。

用法：
    echo '<评测结果 JSON，单条对象或对象数组>' | python3 make_scores_xlsx.py <输出路径.xlsx>

stdin JSON 每项形如：
    {"id":"result_1","name":"策略A","evaluation":{"dimensions":[{"key":...,"name":...,"score":4,"reason":"..."}],"total":3.85}}

输出 .xlsx 列：排名, 条目ID, 条目名称, 相关性, 全面性, 准确性, 可读性, 时效性, 总分
依赖 openpyxl；若缺失，pip install openpyxl 后重试。
"""

import json
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("ERROR: openpyxl 未安装，请先 `pip install openpyxl`", file=sys.stderr)
    sys.exit(1)

DIM_KEYS = ["relevance", "comprehensiveness", "accuracy", "readability", "timeliness"]
DIM_NAMES = {"relevance": "相关性", "comprehensiveness": "全面性", "accuracy": "准确性",
             "readability": "可读性", "timeliness": "时效性"}


def norm(items):
    if isinstance(items, dict):
        return [items]
    if isinstance(items, list):
        return items
    raise ValueError("评测结果需为对象或对象数组")


def rel_score(eval_obj):
    for d in eval_obj["dimensions"]:
        if d["key"] == "relevance":
            return int(d["score"])
    return 0


def main() -> int:
    if len(sys.argv) < 2:
        print("ERROR: 用法: make_scores_xlsx.py <输出.xlsx>", file=sys.stderr)
        return 1
    out_path = sys.argv[1]

    raw = sys.stdin.read().strip()
    if not raw:
        print("ERROR: empty stdin", file=sys.stderr)
        return 1
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as exc:
        print(f"ERROR: invalid JSON: {exc}", file=sys.stderr)
        return 1

    try:
        items = norm(data)
    except ValueError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    # 校验并排序
    valid = []
    for it in items:
        ev = it.get("evaluation") or {}
        dims = ev.get("dimensions") or []
        total = ev.get("total")
        if not (it.get("id") and it.get("name") and dims and isinstance(total, (int, float))):
            continue
        valid.append(it)
    valid.sort(key=lambda r: (-float(r["evaluation"]["total"]), -rel_score(r["evaluation"])))

    wb = Workbook()
    ws = wb.active
    ws.title = "原始打分表"
    header = ["排名", "条目ID", "条目名称", "相关性", "全面性", "准确性", "可读性", "时效性", "总分"]
    ws.append(header)

    header_fill = PatternFill("solid", fgColor="DDEBF7")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    for i, r in enumerate(valid, 1):
        sc = {d["key"]: d["score"] for d in r["evaluation"]["dimensions"]}
        ws.append([
            i, r["id"], r["name"],
            sc.get("relevance"), sc.get("comprehensiveness"), sc.get("accuracy"),
            sc.get("readability"), sc.get("timeliness"),
            round(float(r["evaluation"]["total"]), 2),
        ])

    ws.freeze_panes = "A2"
    widths = [6, 12, 22, 8, 8, 8, 8, 8, 8]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w

    try:
        wb.save(out_path)
    except OSError as exc:
        print(f"ERROR: 保存失败: {exc}", file=sys.stderr)
        return 1
    print(out_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())
