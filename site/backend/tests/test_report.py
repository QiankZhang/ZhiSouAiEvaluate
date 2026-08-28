import io

from openpyxl import load_workbook

from backend import report

MULTI_BM = {
    "eval_method": "MULTI_DIM",
    "config": {
        "dimensions": [
            {"key": "relevance", "name": "相关性", "weight": 50},
            {"key": "quality", "name": "质量", "weight": 50},
        ]
    },
}
GSB_BM = {"eval_method": "GSB", "config": {"dimensions": [{"key": "relevance", "name": "相关性", "weight": 100}]}}

TASK = {
    "id": "TK-9001",
    "name": "测试任务",
    "task_type": "通用评估",
    "benchmark_name": "测试基准",
    "dataset_name": "测试集",
    "judge_model": "gpt-4.1",  # 非 live → 规则聚类，不触发网络
    "progress_total": 3,
    "review_status": "NOT_STARTED",
    "created_at": "2026-08-27 10:00",
    "engine": "simulated",
}


def _multi_result(idx, r_score, q_score, reason):
    total = round((r_score * 50 + q_score * 50) / 100, 2)
    return {
        "row_index": idx,
        "query": f"查询{idx}",
        "content": "内容",
        "status": "SUCCESS",
        "reason": reason,
        "confidence": 0.8,
        "review_status": "PENDING",
        "scores": {
            "dimensions": [
                {"key": "relevance", "name": "相关性", "score": r_score},
                {"key": "quality", "name": "质量", "score": q_score},
            ],
            "total": total,
        },
    }


def test_build_report_multi_dimension_stats_and_lowcount():
    results = [_multi_result(1, 5, 5, "好"), _multi_result(2, 2, 2, "差"), _multi_result(3, 4, 3, "一般")]
    rep = report.build_report(TASK, results, MULTI_BM)
    assert rep["eval_method"] == "MULTI_DIM"
    s = rep["content"]["summary"]
    assert s["total"] == 3
    assert s["low_count"] == 1  # 只有 row2 total=2.0 < 3.0
    dims = {d["key"]: d for d in rep["content"]["dimensions"]}
    assert dims["relevance"]["low_count"] == 1  # row2 relevance=2 <=2
    error_cases = rep["content"]["error_cases"]
    assert error_cases["total"] == 1
    assert error_cases["typical"][0]["query"] == "查询2"


def test_report_markdown_has_five_sections():
    results = [_multi_result(1, 5, 5, "好"), _multi_result(2, 2, 2, "差")]
    rep = report.build_report(TASK, results, MULTI_BM)
    md = report.report_to_markdown(TASK, rep, results)
    for heading in ["## 一、整体结论", "## 三、分维度问题分析", "## 四、典型错误 case 分析", "## 五、改进建议"]:
        assert heading in md


def test_build_report_gsb_net_win_rate():
    def gsb_result(idx, judgment):
        return {
            "row_index": idx,
            "query": f"q{idx}",
            "content": "c",
            "status": "SUCCESS",
            "reason": "r",
            "confidence": 0.8,
            "review_status": "PENDING",
            "scores": {"judgment": judgment, "dimensions": [], "total": 3, "baseline_total": 3},
        }

    results = [gsb_result(1, "Good"), gsb_result(2, "Good"), gsb_result(3, "Bad"), gsb_result(4, "Same")]
    rep = report.build_report({**TASK, "eval_method": "GSB"}, results, GSB_BM)
    s = rep["content"]["summary"]
    assert s["good"] == 2 and s["bad"] == 1 and s["same"] == 1
    assert s["net_win_rate"] == 25.0  # (2-1)/4 * 100


def test_report_xlsx_sheets():
    results = [_multi_result(1, 5, 5, "好"), _multi_result(2, 2, 2, "差")]
    data = report.report_to_xlsx(TASK, results, MULTI_BM)
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["原始打分表"]
    ws = wb["原始打分表"]
    assert ws["A1"].value == "排名"
    assert ws.max_row == 3  # 表头 + 2 行


def test_report_xlsx_gsb_summary_sheet():
    gsb_results = [
        {
            "row_index": 1,
            "query": "q",
            "content": "c",
            "status": "SUCCESS",
            "reason": "r",
            "review_status": "PENDING",
            "scores": {"judgment": "Good", "dimensions": [{"key": "relevance", "name": "相关性", "score": 4}], "total": 4, "baseline_total": 3},
        }
    ]
    data = report.report_to_xlsx({**TASK, "eval_method": "GSB"}, gsb_results, GSB_BM)
    wb = load_workbook(io.BytesIO(data))
    assert "GSB汇总" in wb.sheetnames
