#!/usr/bin/env python3
"""make_scores_xlsx.py — 将多维度评测结果(JSON)转为原始打分表 Excel(.xlsx)，可选附带 GSB 汇总 sheet。

用法：
    echo '<评测结果 JSON>' | python3 make_scores_xlsx.py <输出.xlsx>
    echo '<评测结果 JSON>' | python3 make_scores_xlsx.py <输出.xlsx> <gsb.json 文件路径>

- 评测结果每项形如：
  {"id":"case_1","name":"策略A","evaluation":{"dimensions":[{"key":...,"name":...,"score":4,"reason":"..."}],"total":3.85}}
- GSB JSON 每项形如：
  {"case_id":"case_1","category":"实时性问题","gsb":"G","remark":"..."}

输出 .xlsx：
  sheet1 原始打分表：排名, 条目ID, 条目名称, 相关性, 全面性, 准确性, 可读性, 时效性, 总分
  sheet2 GSB汇总（仅当提供 gsb 文件）：整体/各分类(G/S/B/净胜率) + 逐 case 行
依赖 openpyxl；若缺失，pip install openpyxl 后重试。
"""

import json
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("ERROR: openpyxl 未安装，请先 `pip install openpyxl`", file=sys.stderr)
    sys.exit(1)

DIM_KEYS = ["relevance", "comprehensiveness", "accuracy", "readability", "timeliness"]


def norm(items):
    if isinstance(items, dict):
        return [items]
    if isinstance(items, list):
        return items
    raise ValueError("评测结果需为对象或对象数组")


def rel_score(eval_obj):
    for d in eval_obj["dimensions"]:
        if d["key"] == "relevance":
            return int(d["score"])
    return 0


def net_win_rate(g, s, b):
    denom = g + s + b
    if denom == 0:
        return "—"
    return f"{(g - b) / denom * 100:+.1f}%"


def build_gsb_sheet(wb, gsb_path):
    with open(gsb_path, "r", encoding="utf-8") as f:
        gsb = json.load(f)
    if isinstance(gsb, dict):
        gsb = [gsb]
    if not isinstance(gsb, list):
        raise ValueError("GSB 数据需为对象或对象数组")

    # 聚合：整体 + 按 category
    def counter():
        return {"G": 0, "S": 0, "B": 0}

    overall = counter()
    by_cat = {}
    for item in gsb:
        v = (item.get("gsb") or "").upper()
        if v not in overall:
            continue
        overall[v] += 1
        cat = item.get("category") or "未分类"
        by_cat.setdefault(cat, counter())[v] += 1

    ws = wb.create_sheet("GSB汇总")
    ws.append(["分类", "G", "S", "B", "净胜率"])
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="FCE4D6")
        c.alignment = Alignment(horizontal="center")
    ws.append(["整体", overall["G"], overall["S"], overall["B"],
               net_win_rate(overall["G"], overall["S"], overall["B"])])
    for cat, cnt in sorted(by_cat.items(), key=lambda kv: net_win_rate(kv[1]["G"], kv[1]["S"], kv[1]["B"])):
        ws.append([cat, cnt["G"], cnt["S"], cnt["B"], net_win_rate(cnt["G"], cnt["S"], cnt["B"])])
    ws.append([])
    ws.append(["逐 case", "case_id", "category", "gsb", "remark"])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True, italic=True)
    for item in gsb:
        ws.append(["", item.get("case_id", ""), item.get("category", ""),
                   (item.get("gsb") or "").upper(), item.get("remark", "")])
    for col, w in zip("ABCDE", [16, 12, 14, 8, 48]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


def main() -> int:
    if len(sys.argv) < 2:
        print("ERROR: 用法: make_scores_xlsx.py <输出.xlsx> [gsb.json]", file=sys.stderr)
        return 1
    out_path = sys.argv[1]
    gsb_path = sys.argv[2] if len(sys.argv) > 2 else None

    raw = sys.stdin.read().strip()
    if not raw:
        print("ERROR: empty stdin", file=sys.stderr)
        return 1
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as exc:
        print(f"ERROR: invalid JSON: {exc}", file=sys.stderr)
        return 1

    try:
        items = norm(data)
    except ValueError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    valid = []
    for it in items:
        ev = it.get("evaluation") or {}
        dims = ev.get("dimensions") or []
        total = ev.get("total")
        if not (it.get("id") and it.get("name") and dims and isinstance(total, (int, float))):
            continue
        valid.append(it)
    valid.sort(key=lambda r: (-float(r["evaluation"]["total"]), -rel_score(r["evaluation"])))

    wb = Workbook()
    ws = wb.active
    ws.title = "原始打分表"
    ws.append(["排名", "条目ID", "条目名称", "相关性", "全面性", "准确性", "可读性", "时效性", "总分"])
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDEBF7")
        c.alignment = Alignment(horizontal="center")

    for i, r in enumerate(valid, 1):
        sc = {d["key"]: d["score"] for d in r["evaluation"]["dimensions"]}
        ws.append([
            i, r["id"], r["name"],
            sc.get("relevance"), sc.get("comprehensiveness"), sc.get("accuracy"),
            sc.get("readability"), sc.get("timeliness"),
            round(float(r["evaluation"]["total"]), 2),
        ])

    ws.freeze_panes = "A2"
    for idx, w in enumerate([6, 12, 22, 8, 8, 8, 8, 8, 8], 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w

    if gsb_path:
        try:
            build_gsb_sheet(wb, gsb_path)
        except (OSError, ValueError, json.JSONDecodeError) as exc:
            print(f"ERROR: GSB sheet 构建失败: {exc}", file=sys.stderr)
            return 1

    try:
        wb.save(out_path)
    except OSError as exc:
        print(f"ERROR: 保存失败: {exc}", file=sys.stderr)
        return 1
    print(out_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())
